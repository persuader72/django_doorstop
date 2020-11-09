from typing import Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.fill import SolidColorFillProperties

from doorstop import Tree, Item
from doorstop.core.document import Document


def import_from_xslx(doc):
    #  type: (Document) -> None
    fields = ['uid', 'header', 'text', 'level', 'parent', 'subsystem']
    field_indexes = [-1, -1, -1, -1, -1, -1]

    def get_column_field(field, _ws, row):
        #  type: (str, Worksheet, int) -> Optional[str]
        _i = fields.index(field)
        _ii = field_indexes[_i]
        c = _ws.cell(row, _ii)
        if c is None or c.value is None:
            return None
        stripped = c.value.strip()
        if len(stripped) == 0:
            return None
        return stripped

    xlsxfile = '/tmp/import.xlsx'
    wb = load_workbook(xlsxfile)
    ws = wb.active

    keys = []
    cols = ws.max_column
    rows = ws.max_row
    for i in range(1, cols+1):
        keys.append(ws.cell(1, i).value.strip())
    for i in range(0, len(fields)):
        field_indexes[i] = keys.index(fields[i])+1

    for i in range(2, rows+1):
        uid = get_column_field('uid', ws, i)
        if uid is None:
            item = doc.add_item()
        else:
            item = doc.find_item(uid)

        item.header = get_column_field('header',ws, i)
        item.text = get_column_field('text',ws, i)
        level = get_column_field('level', ws, i)
        if level is not None:
            item.level = level
        parent = get_column_field('parent', ws, i)
        if parent is not None:
            pdoc = doc.tree.find_document('RADN')  # type: Document
            for pitem in pdoc.items:
                if pitem.get('orig_ref') == parent:
                    print(item.uid)
                    doc.tree.link_items(item.uid, pitem.uid)
        item.set('subsystem', get_column_field('subsystem',ws, i))


def export_doc_to_xlsx(doc, ws):
    # type: (Document, Worksheet) -> None
    ws.title = doc.prefix
    ws.cell(1, 1, "uid")
    ws.cell(1, 2, "header")
    ws.cell(1, 3, "text")
    ws.cell(1, 4, "level")
    ws.cell(1, 5, "pending")
    ws.cell(1, 6, "deleted")
    ws.cell(1, 7, "parent")

    ws.column_dimensions['A'].width *= 1.5
    ws.column_dimensions['B'].width *= 4
    ws.column_dimensions['C'].width *= 8

    for i, ff in enumerate(doc.forgein_fields):
        ws.cell(1, 8+i, ff)
    ws.cell(1, 8+len(doc.forgein_fields), "comments")
    ws.column_dimensions[get_column_letter(8+len(doc.forgein_fields))].width *= 12

    header_font = Font(color='FFFFFFFF')
    header_fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
    for i in range(1, 8+len(doc.forgein_fields)+1):
        ws.cell(1,i).fill = header_fill
        ws.cell(1,i).font = header_font

    last_row = 2
    for item in doc.items:
        ws.cell(last_row, 1, str(item.uid))
        if item.deleted:
            ws.cell(last_row, 1).font = Font(strike=True)
        ws.cell(last_row, 2, item.header)
        if item.deleted:
            ws.cell(last_row, 2).font = Font(strike=True)
        ws.cell(last_row, 3, item.text)
        if item.deleted:
            ws.cell(last_row, 3).font = Font(strike=True)
        ws.cell(last_row, 4, str(item.level))
        pitems = [x.uid.value for x in item.parent_items]
        ws.cell(last_row, 5, "X" if item.pending else "")
        ws.cell(last_row, 6, "X" if item.deleted else "")
        ws.cell(last_row, 7, str(','.join(pitems)))

        for i, ff in enumerate(doc.forgein_fields):
            ffi = doc.forgein_fields[ff]
            if ffi['type'] == 'multi':
                ws.cell(last_row, 8+i, ','.join(item.get(ff)))
            else:
                ws.cell(last_row, 8+i, item.get(ff))

        i = 8 + len(doc.forgein_fields)
        comments = item.get('comments')
        if comments is not None:
            text = ''
            for comment in comments:
                if len(text) > 0:
                    text += '\n'
                text += '[X] ' if 'closed' in  comment and comment['closed'] is True else ' [ ]'
                text += comment['text']
                text += f' {comment["author"]} {comment["date"]}'
            ws.cell(last_row, i, text)
            ws.cell(last_row, i).alignment = Alignment(wrapText=True)

        last_row += 1

def export_doc_treac(parent_doc, child_doc, ws):
    #  type: (Document, Document, Worksheet) -> None
    ws.title = f'{parent_doc.prefix} vs {child_doc.prefix}'
    ws.cell(1, 1, f'req of {parent_doc.prefix}')
    ws.cell(1, 2, f'req of {child_doc.prefix}')

    ws.column_dimensions['A'].width *= 4
    ws.column_dimensions['B'].width *= 8

    header_font = Font(color='FFFFFFFF')
    header_fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
    for i in range(1, 2):
        ws.cell(1,i).fill = header_fill
        ws.cell(1,i).font = header_font

    last_row = 2
    for item in parent_doc.items:
        if item.normative:
            ws.cell(last_row, 1, str(item.uid))
            childs = ''
            for child in item.find_child_items():  # type: (Item)
                if child.document.prefix == child_doc.prefix:
                    childs += str(child.uid) if len(childs) == 0 else ', ' + str(child.uid)
            ws.cell(last_row, 2, childs)
            last_row += 1


def export_doc_reverse_treac(parent_doc, child_doc, ws):
    #  type: (Document, Document, Worksheet) -> None
    ws.title = f' {child_doc.prefix} vs {parent_doc.prefix}'
    ws.cell(1, 1, f'req of {child_doc.prefix}')
    ws.cell(1, 2, f'req of {parent_doc.prefix}')

    ws.column_dimensions['A'].width *= 4
    ws.column_dimensions['B'].width *= 8

    header_font = Font(color='FFFFFFFF')
    header_fill = PatternFill(start_color='FF000000', end_color='FF000000', fill_type='solid')
    for i in range(1, 2):
        ws.cell(1,i).fill = header_fill
        ws.cell(1,i).font = header_font

    last_row = 2
    for item in child_doc.items:
        if item.normative:
            ws.cell(last_row, 1, str(item.uid))
            parents = ''
            for _parent in item.parent_items:  # type: (Item)
                if _parent.document.prefix == parent_doc.prefix:
                    parents += str(_parent.uid) if len(parents) == 0 else ', ' + str(_parent.uid)
            ws.cell(last_row, 2, parents)
            last_row += 1


def export_to_xlsx(doc):
    xlsxfile = '/tmp/export.xlsx'
    wb = Workbook()
    ws = wb.active
    export_doc_to_xlsx(doc, ws)
    wb.save(xlsxfile)
    return xlsxfile


def export_full_xslx(tree):
    #  type: (Tree) -> str
    wb = Workbook()
    for i, doc in enumerate(tree.documents):
        if i >= len(wb.worksheets):
            wb.create_sheet()
        ws = wb.worksheets[i]
        export_doc_to_xlsx(doc, ws)
    for i, doc in enumerate(tree.documents):  # type: (int, Document)
        childs = []
        for j, _child in enumerate(tree.documents):  # type: (int, Document)
            if _child.parent == doc.prefix:
                childs.append(_child)
        for _child in childs:
            ws = wb.create_sheet()
            export_doc_treac(doc, _child, ws)
            ws = wb.create_sheet()
            export_doc_reverse_treac(doc, _child, ws)
    xlsxfile = '/tmp/export.xlsx'
    wb.save(xlsxfile)
    return xlsxfile
